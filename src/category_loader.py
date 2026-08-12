# -*- coding: utf-8 -*-
"""商品カテゴリー設定ファイル（Excel）ローダーモジュール。

config.md.xlsx（または config.xlsx, config.md）内の全シートを読み込み、
シート名（カテゴリー名）と各シート内の商品コード・商品名から
高速なマッピングテーブルを構築して商品へのカテゴリ付与を行います。
"""

import os
import glob
import logging
import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)

# 設定ファイルの候補パス
DEFAULT_CONFIG_CANDIDATES = [
    "config.md.xlsx",
    "config.xlsx",
    "config.md",
    "input_data/config.md.xlsx",
    "input_data/config.xlsx",
]


def normalize_code(code_val) -> str:
    """商品コードを検索用文字列に正規化する（小数点除去・13桁ゼロ埋め）。"""
    if code_val is None or pd.isna(code_val):
        return ""
    s = str(code_val).strip()
    if "." in s:
        s = s.split(".")[0]
    if s.isdigit():
        return s.zfill(13)
    return s


def normalize_name(name_val) -> str:
    """商品名を検索用文字列に正規化する。"""
    if name_val is None or pd.isna(name_val):
        return ""
    return str(name_val).strip()


class CategoryMapper:
    """カテゴリー設定ファイルからマッピングテーブルを管理するクラス。"""

    def __init__(self, config_path: str | None = None):
        self.config_path = config_path or self._find_config_file()
        self.code_to_category: dict[str, str] = {}
        self.name_to_category: dict[str, str] = {}
        self.category_list: list[str] = []
        self._loaded = False

        if self.config_path and os.path.exists(self.config_path):
            self.load()
        else:
            logger.warning(
                "カテゴリー設定ファイルが見つかりませんでした。デフォルトの「その他」分類が適用されます。"
            )

    def _find_config_file(self) -> str | None:
        """設定ファイルパスを自動探索する。"""
        # 1. 定義済み候補パスの確認
        for path in DEFAULT_CONFIG_CANDIDATES:
            if os.path.isfile(path):
                return path

        # 2. ワイルドカード探索 (*config*.xlsx)
        for pattern in ["*config*.xlsx", "input_data/*config*.xlsx"]:
            matches = glob.glob(pattern)
            if matches:
                return matches[0]

        return None

    def load(self) -> None:
        """Excel設定ファイルを読み込み、マッピングテーブルを構築する。"""
        if not self.config_path or not os.path.isfile(self.config_path):
            logger.error("設定ファイルが存在しません: %s", self.config_path)
            return

        logger.info("カテゴリー設定ファイルを読み込み中: %s", self.config_path)
        try:
            wb = openpyxl.load_workbook(self.config_path, read_only=True, data_only=True)
        except Exception as e:
            logger.error("設定ファイル読み込みエラー (%s): %s", self.config_path, e)
            return

        code_map = {}
        name_map = {}
        categories = []

        for sheet_name in wb.sheetnames:
            clean_category_name = sheet_name.strip()
            if not clean_category_name:
                continue

            categories.append(clean_category_name)
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # ヘッダー行とカラム位置の特定
            header_idx = -1
            code_col = 0
            name_col = 1
            for idx, row in enumerate(rows[:5]):
                row_str = [str(c).strip() if c is not None else "" for c in row]
                if "商品コード" in row_str:
                    header_idx = idx
                    code_col = row_str.index("商品コード")
                    name_col = 1 if code_col == 0 else 0
                    if "商品名" in row_str:
                        name_col = row_str.index("商品名")
                    break

            data_rows = rows[header_idx + 1 :] if header_idx != -1 else rows

            for r in data_rows:
                if not r:
                    continue
                raw_code = r[code_col] if len(r) > code_col else None
                raw_name = r[name_col] if len(r) > name_col else None

                code_str = normalize_code(raw_code)
                name_str = normalize_name(raw_name)

                if not code_str and not name_str:
                    continue
                if code_str == "商品コード" or name_str == "商品名":
                    continue

                if code_str:
                    code_map[code_str] = clean_category_name
                if name_str:
                    name_map[name_str] = clean_category_name

        wb.close()

        self.code_to_category = code_map
        self.name_to_category = name_map
        self.category_list = categories
        self._loaded = True

        logger.info(
            "カテゴリー設定の読み込み完了: カテゴリ数=%d, コード数=%d, 商品名数=%d",
            len(self.category_list),
            len(self.code_to_category),
            len(self.name_to_category),
        )

    def get_category(self, code_val, name_val=None, default: str = "その他") -> str:
        """商品コードまたは商品名からカテゴリー名を取得する。"""
        # 1. 商品コードでの完全一致判定
        code_str = normalize_code(code_val)
        if code_str and code_str in self.code_to_category:
            return self.code_to_category[code_str]

        # 2. 商品名での完全一致判定
        if name_val is not None:
            name_str = normalize_name(name_val)
            if name_str and name_str in self.name_to_category:
                return self.name_to_category[name_str]

        return default

    def apply_to_dataframe(
        self,
        df: pd.DataFrame,
        code_col: str = "code",
        name_col: str = "name",
        target_col: str = "category",
        default: str = "その他",
    ) -> pd.DataFrame:
        """DataFrameにカテゴリー列を追加/更新する。"""
        if df.empty:
            df[target_col] = pd.Series(dtype="str")
            return df

        has_code = code_col in df.columns
        has_name = name_col in df.columns

        if has_code and has_name:
            categories = [
                self.get_category(row_code, row_name, default=default)
                for row_code, row_name in zip(df[code_col], df[name_col])
            ]
        elif has_code:
            categories = [
                self.get_category(row_code, default=default)
                for row_code in df[code_col]
            ]
        elif has_name:
            categories = [
                self.get_category(None, row_name, default=default)
                for row_name in df[name_col]
            ]
        else:
            categories = [default] * len(df)

        df[target_col] = categories
        return df


# シングルトンインスタンスのキャッシュ管理
_GLOBAL_MAPPER: CategoryMapper | None = None


def get_category_mapper(config_path: str | None = None) -> CategoryMapper:
    """CategoryMapperのグローバルインスタンスを取得する。"""
    global _GLOBAL_MAPPER
    if _GLOBAL_MAPPER is None or config_path is not None:
        _GLOBAL_MAPPER = CategoryMapper(config_path=config_path)
    return _GLOBAL_MAPPER


def apply_categories(
    df_week: pd.DataFrame,
    df_day: pd.DataFrame,
    config_path: str | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """2つのDataFrameに対して一括でカテゴリ分類を適用する。"""
    mapper = get_category_mapper(config_path=config_path)
    df_week = mapper.apply_to_dataframe(df_week)
    df_day = mapper.apply_to_dataframe(df_day)
    return df_week, df_day


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    mapper = get_category_mapper()
    print("=== Category Mapper Status ===")
    print(f"Config path: {mapper.config_path}")
    print(f"Loaded sheets/categories: {len(mapper.category_list)}")
    print(f"Sample categories: {mapper.category_list[:10]}")
    print(f"Total codes mapped: {len(mapper.code_to_category)}")
    print(f"Total names mapped: {len(mapper.name_to_category)}")
